from rest_framework import views
from rest_framework.response import Response
from rest_framework import status

from django.http import HttpResponse
from django.db.models import Q
from django.core.exceptions import ObjectDoesNotExist

import openpyxl

from academics import models as academics_models

from event import constants as event_constants

from . import models as account_models
from . import constants as account_constants
from . import permissions


class ClassroomExport(views.APIView):

    """
    Generate an excel report for a given classroom,
    with separate sheets for academics and events.
    """
    permission_classes = [permissions.IsStaff]
    def get(self,request,classroom_id):
        """
        when a get request is given to this view
        it takes classroom_id from the url and returns
        an excel sheet of the class details about academics 
        and events
        """
        classroom = academics_models.Classroom.objects.get(id=classroom_id)
        class_teacher = classroom.teacher
        
        if request.user.account.user_type == account_constants.UserType.TEACHER:
            if class_teacher == request.user.account.teacher:
                pass
            else:
                return Response(
                    {"message":"this teacher is not class teacher of \
                    the student"},status=status.HTTP_401_UNAUTHORIZED)
            
        # Parse the sheets parameter
        sheets_param = request.query_params.get('sheets')
        if sheets_param:
            sheets = sheets_param.split(',')
        else:
            sheets = ['academic', 'event']
        
        # create a new workbook
        wb = openpyxl.Workbook()

        if 'academic' in sheets:
            #creating a sheet
            academic_sheet = wb.active
            academic_sheet.title = 'Academics'

            # add headers
            academic_sheet.append(["Exam Name"," Start Time","End Time"])
            exams = classroom.exams.all()
            
            for exam in exams:
                name = exam.name
                start = str(exam.start_time)
                end = str(exam.end_time)
                academic_sheet.append([name,start,end])
                
        if 'event' in sheets:
        # Create a sheet for events
            event_sheets = wb.create_sheet("Events")
            event_sheets.append(["Event Name","Fest Name","Event Type"])
            events = classroom.events.all()

            for event in events:
                name = event.name
                fest = event.fest.name
                event_type = event.event_type
                if event_type == event_constants.EventType.TIME:
                    event_type = 'RACE'
                else:
                    event_type = 'THROW'

                event_sheets.append([name,fest,event_type])
    
        # creates an HttpResponse object with the content type excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument. \
            spreadsheetml.sheet')
        # set filename 
        response['Content-Disposition'] = f'attachment; \
        filename=classreport.xlsx'

        wb.save(response)
        return response


class StudentReportExport(views.APIView):
    """
    Generate an excel report for a given student
    the detail includes his performance in exams
    """
    permission_classes = [permissions.IsStaff]

    def get(self, request,student_id):
        """
        get request accepts student_id from url
        and returns the performance of the student in academics 
        as excel with the ddtails of the student
        """
        student = account_models.Student.objects.get(user__user__id=student_id)

        # selects latest enrollment done by the student        
        latest_enrollment = student.enrollment_set.order_by(
            '-enroll_date').first()
        full_name = student.user.get_fullname()
        classroom = latest_enrollment.classroom
        standard = classroom.standard
        division = classroom.division
        class_teacher = classroom.teacher

        if request.user.account.account_type == account_constants.UserType.TEACHER:
             if class_teacher == request.user.account.teacher:
                  pass
             else:
                  return Response(
                       {"message":"this teacher is not class teacher of \
                        the student"},status=status.HTTP_401_UNAUTHORIZED)
        
        # open a workbook
        wb = openpyxl.Workbook()
        student_sheet = wb.active
        student_sheet.title = 'Academics'

        a4 = student_sheet['A4']
        b4 = student_sheet['B4']
        c4 = student_sheet['C4']

        # stores fill colors to fill in cells
        fill1 = openpyxl.styles.PatternFill(
            start_color='76332F', end_color='76332F', fill_type='solid')
        fill2 = openpyxl.styles.PatternFill(
            start_color='F0D020', end_color='F0D020', fill_type='solid')
        
        # set the student details in excel
        student_sheet['A1'] = 'Full Name'
        student_sheet['B1'] = full_name
        student_sheet['A2'] = 'Standard'
        student_sheet['B2'] = standard
        student_sheet['A3'] = 'Division'
        student_sheet['B3'] = division
        student_sheet['A4'] = 'Exam'
        student_sheet['B4'] = 'Max Mark'
        student_sheet['C4'] = 'Score'

        # Apply the colors to the cells
        student_sheet['A1'].fill = fill1
        student_sheet['A2'].fill = fill1
        student_sheet['A3'].fill = fill1

        for cell in [a4,b4,c4]:
                    cell.fill = fill2    
        
        # takes all exam object
        exams = latest_enrollment.classroom.exams.all()
        for exam in exams:
            questions = academics_models.Question.objects.filter(exam=exam)
            mark = 0
            for question in questions:
                mark = mark + question.mark
            total_mark = mark # max mark that can be scored
            responses = academics_models.Response.objects.filter(
                student=student,option__question__exam =exam)
            if responses:
                score =0
                for response in responses:
                    score = score + response.option.filter(
                        is_correct=True).count()
                score = score #mark of the student
                # store data to sheet 
                student_sheet.append([exam.name,total_mark,score])
            else:
                student_sheet.append([exam.name,total_mark,"Not Written"])

        # creates an HttpResponse object with the content type excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument \
            .spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; \
        filename=studentreport.xlsx'

        wb.save(response)
        return response
    

class SchoolMembers(views.APIView):
    """
    Generate an excel report with every members of the 
    school

    permission
    - Only admins can give this request
    """
    permission_classes = [permissions.IsAdmin]

    def get(self, request):
        """
        this  get request returns all members of the school
        with their classroom details and role
        """
        teachers = account_models.Teacher.objects.all()
        students = account_models.Student.objects.all()
        admins = account_models.Account.objects.filter(
            user_type = account_constants.UserType.ADMIN)
        
        wb = openpyxl.Workbook()
        members_sheet = wb.active

        fill1 = openpyxl.styles.PatternFill(
            start_color='76332F', end_color='76332F', fill_type='solid')

        # Set the headers
        members_sheet['A1'] = 'Full Name'
        members_sheet['B1'] = 'Role'
        members_sheet['C1'] = 'Standard'
        members_sheet['D1'] = 'Division'

        # Apply the colors to the cells
        members_sheet['A1'].fill = fill1
        members_sheet['B1'].fill = fill1
        members_sheet['C1'].fill = fill1      
        members_sheet['D1'].fill = fill1      

        # add teacher data to sheet
        for teacher in teachers:
            full_name = teacher.user.get_fullname()
            role = 'TEACHER'
            try :
                standard = teacher.classroom.standard
                division = teacher.classroom.division
            except ObjectDoesNotExist:
                standard = None
                division = None
            members_sheet.append([full_name,role,standard,division])
        
        # add student data to sheet
        for student in students:
            full_name = teacher.user.get_fullname()
            role = 'STUDENT'
            latest_enrollment = student.enrollment_set.order_by(
                '-enroll_date').first()
            try:
                standard = latest_enrollment.classroom.standard
                division = latest_enrollment.classroom.division
            except:
                standard = None
                division = None
            members_sheet.append([full_name,role,standard,division])
               
        # add admin data to sheet
        for admin in admins:
            full_name = admin.get_fullname()
            role = 'ADMIN'
            members_sheet.append([full_name,role])
        
        # creates an HttpResponse object with the content type excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=schoolreport.xlsx'

        wb.save(response)
        return response


class BestTeacher(views.APIView):
    """
    Generate an excel which shows best performing 
    teacher in order
    permission
    - Only admins can give this request
    """
    permission_classes = [permissions.IsAdmin]

    def get(self, request):
        """
        this get request returns a list of teacher along with their 
        performance in percentage
        """
        classrooms = academics_models.Classroom.objects.all()
        # create a dictionary to store the total marks of each teacher
        teacher_performance = {}

        wb = openpyxl.Workbook()
        teacher_sheet = wb.active

        fill1 = openpyxl.styles.PatternFill(
            start_color='76332F', end_color='76332F', fill_type='solid')
        
        # set header in sheets
        teacher_sheet['A1'] = 'Full Name'
        teacher_sheet['B1'] = 'Performance'

        # set color to header cells
        teacher_sheet['A1'].fill = fill1
        teacher_sheet['B1'].fill = fill1

        for classroom in classrooms:
            students = classroom.students.all()
            # total number of questions in latest exam
            try:
                total_questions = classroom.exams.latest(
                    'start_time').questions.count()
                exam = classroom.exams.latest('start_time')
            except ObjectDoesNotExist:
                exam = None
                total_questions = None

            # get the total number of correct responses
            correct_responses = academics_models.Response.objects.filter(
                option__is_correct=True,
                option__question__exam=exam,
                student__in=students,
            ).count()

            try:
                teacher_performance[classroom.teacher.user.user.username] = (
                    correct_responses / (total_questions * students.count())) * 100
            except Exception as e:
                teacher_performance[classroom.teacher.user.user.username]=0

        # sort the teacher performance   
        teacher_performance = dict(sorted(teacher_performance.items(), key=lambda x: x[1], reverse=True))

        # enter teacher details into sheets
        for key,value in teacher_performance.items():
            teacher_sheet.append([key,value])
        
        # creates an HttpResponse object with the content type excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=teacherreport.xlsx'

        wb.save(response)
        return response


        
        



